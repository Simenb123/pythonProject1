# utvalg_logikk.py – 2025-05-18  Oppsummering v4 (diagrammer)
from __future__ import annotations
from pathlib import Path
from typing  import Mapping
import pandas as pd, re, unicodedata, datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference

_KR   = u'_-* # ##0\\ kr_-;\\-* -# ##0\\ kr_-;_-* "-"??\\ kr_-;_-@_-'
_PCT  = u'0,0 %'
_INT  = u'# ##0'
_DATE = "DD.MM.YYYY TT:MM"

# ---------- hjelpere (samme som før – forkortet her) ----------------------
def _norm(t:str)->str:
    return re.sub(r"[^0-9a-z]","",unicodedata.normalize("NFKD",t)
                  .encode("ascii","ignore").decode().lower())

def _les(src:Path, meta:Mapping[str,str])->pd.DataFrame:
    if (pq:=meta.get("std_file")) and Path(pq).exists():
        return pd.read_parquet(pq, engine="pyarrow")
    if src.suffix.lower() in (".xlsx",".xls"):
        return pd.read_excel(src, engine="openpyxl")
    kw={"sep":None,"engine":"python"}
    if meta.get("encoding"): kw["encoding"]=meta["encoding"]
    return pd.read_csv(src, **kw)

def _apply(df:pd.DataFrame, mapping:Mapping[str,str])->pd.DataFrame:
    n2c={_norm(c):c for c in df.columns}
    for std,raw in mapping.items():
        if std in ("encoding","std_file"): continue
        w=_norm(raw)
        if w in n2c: df=df.rename(columns={n2c[w]:std})
    for std,raw in mapping.items():
        if std in ("encoding","std_file") or std in df.columns: continue
        cand=next((c for c in df.columns if _norm(std) in _norm(c)),None)
        if cand is None and raw in df.columns: cand=raw
        if cand is not None: df[std]=df[cand]
    return df

def _auto_fmt(ws):
    hdr_fill=PatternFill("solid","D9D9D9")
    hdr_font=Font(bold=True)
    hdr_align=Alignment(horizontal="center")
    for idx,col in enumerate(ws.columns,1):
        c0=col[0]; c0.fill=hdr_fill; c0.font=hdr_font; c0.alignment=hdr_align
        hdr=(c0.value or "").lower()
        fmt=_INT
        if any(t in hdr for t in ("beløp","sum")): fmt=_KR
        elif "%" in hdr:                          fmt=_PCT
        elif "dato" in hdr or "tidspunkt" in hdr: fmt=_DATE
        for c in col[1:]: c.number_format=fmt
        w=max([len(str(c.value or "")) for c in col[:200]]+[8])*1.2
        ws.column_dimensions[get_column_letter(idx)].width=w

# ---------- hovedfunksjon -------------------------------------------------
def kjør_bilagsuttrekk(
    src:Path,
    konto_rng:tuple[int,int],
    belop_rng:tuple[float,float],
    n_bilag:int,
    *,
    meta:Mapping[str,str],
)->dict[str,Path]:

    df=_apply(_les(src,meta),meta)
    k,b,bn="konto","beløp","bilagsnr"
    for col in (k,b,bn,"dato"):
        if col not in df.columns: raise ValueError(f"Mangler «{col}»")

    df[k]=(df[k].astype(str).str.replace(r"\D","",regex=True)
                   .astype("Int32",errors="ignore"))
    filt=df[df[k].between(*konto_rng)&df[b].between(*belop_rng)]
    unike=filt[bn].drop_duplicates()
    if n_bilag>len(unike):
        raise ValueError(f"Filtrert pop har bare {len(unike)} bilag")
    valgte=unike.sample(n=n_bilag).sort_values()

    fullt    = df[df[bn].isin(valgte)]
    interval = filt[filt[bn].isin(valgte)]
    summer   =(interval.groupby(bn)[b]
               .agg(Sum_i_intervallet="sum",
                    Linjer_i_intervallet="count").reset_index())

    pop_sum, uttrekk_sum = filt[b].sum(), interval[b].sum()

    opp=pd.DataFrame({
        "Beskrivelse":[
          "Populasjon – unike bilagsnr",
          "Populasjon – sum beløp",
          "Utvalgsstørrelse (bilag)",
          "Uttrekket dekker (%)",
          "Uttrekk – sum beløp",
          "Uttrekk sum (% av pop.)",
          "Konto-intervall","Beløps-intervall","Tidspunkt"],
        "Verdi":[
          len(unike), pop_sum, n_bilag, n_bilag/len(unike),
          uttrekk_sum, uttrekk_sum/pop_sum if pop_sum else 0,
          f"{konto_rng[0]} – {konto_rng[1]}",
          f"{belop_rng[0]:,.0f} – {belop_rng[1]:,.0f}",
          dt.datetime.now().strftime("%Y-%m-%d %H:%M")]})

    # ---- rekkefølge i detaljfaner --------------------
    order=[std for std in meta if std not in ("encoding","std_file")
           and std in fullt.columns]
    _r=lambda fr: fr[order+[c for c in fr.columns if c not in order]]
    fullt,interval=map(_r,(fullt,interval))

    # ---- filnavn m/ versjonering ---------------------
    base=src.parent/f"Bilag_uttrekk_{n_bilag}.xlsx"
    out=base; i=2
    while out.exists():
        out=base.with_stem(f"{base.stem}_v{i}"); i+=1

    # ---- skriv Excel ---------------------------------
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        opp.to_excel     (xw,"Oppsummering",      index=False)
        summer.to_excel  (xw,"Bilag_summer",      index=False)
        interval.to_excel(xw,"Kun_intervallet",   index=False)
        fullt.to_excel   (xw,"Fullt_bilagsutvalg",index=False)

    # ---- pynt & diagrammer ---------------------------
    wb=load_workbook(out)
    for ws in wb.worksheets: _auto_fmt(ws)

    ws_opp=wb["Oppsummering"]

    # Hyperlenker
    for r,(sheet,txt) in enumerate(
        [("Bilag_summer","Bilag-summer"),
         ("Kun_intervallet","Kun-intervallet"),
         ("Fullt_bilagsutvalg","Fullt-bilagsutvalg")], 12):
        c=ws_opp.cell(row=r, column=1, value=txt)
        c.hyperlink=f"#{sheet}!A1"; c.style="Hyperlink"

    # Sektor­diagram – beløpsdekning
    pie = PieChart()
    pie.add_data(Reference(ws_opp, min_col=2, min_row=4, max_row=5), titles_from_data=False)
    pie.set_categories(Reference(ws_opp, min_col=1, min_row=4, max_row=5))
    pie.title="Uttrekk vs Populasjon (beløp)"
    ws_opp.add_chart(pie,"D2")

    # Histogram – 20 kvantiler
    q=filt[b].quantile([i/20 for i in range(21)]).reset_index()
    for i,(kval,val) in enumerate(q.values,1):
        ws_opp.cell(row=i, column=6, value=f"P{int(kval*100):>3}")
        ws_opp.cell(row=i, column=7, value=val).number_format=_KR
    bar=BarChart()
    bar.add_data(Reference(ws_opp,min_col=7,min_row=1,max_row=21),
                 titles_from_data=False)
    bar.set_categories(Reference(ws_opp,min_col=6,min_row=1,max_row=21))
    bar.title="Beløpsfordeling (20-kvantiler)"
    bar.height,bar.width=6,13
    ws_opp.add_chart(bar,"D18")

    wb.save(out)
    return {"uttrekk":out,"valgte_bilag":valgte.tolist()}
