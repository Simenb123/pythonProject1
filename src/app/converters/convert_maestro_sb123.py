# convert_maestro_sb123.py
# ---------------------------------------------------------
# Genererer “Data”, “Pivot”, “KontoPivot”, “Oppstilling” og
# “Vesentlighetsgrense”.  Rapporten lagres i *samme* mappe som
# kildefilen med navnet  <original-filnavn> Start[.vN].xlsx
#
# NYTT 2025-05-XX
#   • load_source() håndterer kontonr med punktum/strekk («1550.1» …).
#   • Husker sist brukte mappe (.last_dir.json).
#   • KontoPivot inneholder nå Regnnr / Regnskapslinje / NA-konto / NA-navn.
# ---------------------------------------------------------
from __future__ import annotations
import argparse, json, logging, re, sys
from pathlib import Path
from tkinter import Tk, filedialog
from typing import List, Optional

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from src.app.services.oppstilling import lag_oppstilling

# -------------------------------------------------- #
#  konfig
# -------------------------------------------------- #
BASE_DIR_DEF = Path(r"F:/Dokument/2/SB")
MAESTRO_FILE  = "RES OG BAL MAESTRO.xlsx"
NAERING_FILE  = "Naeringsspesifikasjon.xlsx"
REVISJON_FILE = "Revisjonsområder.xlsx"
AAR = "2024"                       # brukt i Vesentlighets-fanen
pd.options.mode.copy_on_write = True

# -------------------------------------------------- #
#  persistert “sist brukte mappe”
# -------------------------------------------------- #
STATE_FILE = Path(__file__).with_suffix(".last_dir.json")
def _read_last_dir() -> Optional[Path]:
    try:
        return Path(json.loads(STATE_FILE.read_text())["last"])
    except Exception:
        return None
def _write_last_dir(p: Path) -> None:
    try:
        STATE_FILE.write_text(json.dumps({"last": str(p)}, indent=2))
    except Exception:
        pass

# -------------------------------------------------- #
#  dialog
# -------------------------------------------------- #
def choose_source(fallback: Path) -> Path:
    start_dir = _read_last_dir() or fallback
    root = Tk(); root.withdraw()
    fname = filedialog.askopenfilename(
        title="Velg Saldobalanse/Kontoplan fra Maestro",
        initialdir=str(start_dir),
        filetypes=[("Excel-filer", "*.xlsx *.xls *.xlsm")])
    root.destroy()
    if not fname:
        sys.exit("Ingen fil valgt – prosessen avbrytes.")
    p = Path(fname)
    _write_last_dir(p.parent)
    return p

# -------------------------------------------------- #
#  les referansetabell
# -------------------------------------------------- #
def read_ref(xl: Path, *, key_col: int, val_col: int, sheet=0):
    df = pd.read_excel(xl, sheet_name=sheet, header=None,
                       usecols=[key_col, val_col], engine="openpyxl")
    df.columns = ["Key", "Value"]
    df["Key"] = pd.to_numeric(df["Key"], errors="coerce")
    df.dropna(subset=["Key"], inplace=True)
    if df.empty:
        raise ValueError("ingen rader")
    df["Key"] = df["Key"].astype("Int64").astype(str)
    return df.set_index("Key")

# -------------------------------------------------- #
#  kildefil → Data-frame
# -------------------------------------------------- #
def _clean_acct(v) -> str:
    """Beholder bare 0-9 samt . og - (fjerner annet støy)."""
    if pd.isna(v):
        return ""
    s = str(v).strip().replace(" ", "")
    m = re.match(r"[\d\.\-]+", s)
    return m.group(0) if m else ""

def load_source(xl: Path) -> pd.DataFrame:
    raw = pd.read_excel(xl, header=None, skiprows=3, engine="openpyxl")
    col_map = {0:"Konto",1:"Kontonavn",6:"Saldo i fjor",
               3:"Foreløpig Saldo i år",4:"Korreksjon i år",5:"Saldo i år",
               7:"H",8:"I",12:"NAkonto",15:"Revnr"}
    df = raw[list(col_map)].rename(columns=col_map)

    # kontonummer – støtter «1550.1» osv.
    df = df[~df["Konto"].isna()].copy()
    df["Konto"] = df["Konto"].apply(_clean_acct)
    mask_int = df["Konto"].str.fullmatch(r"\d+")
    df.loc[mask_int, "Konto"] = df.loc[mask_int, "Konto"].astype("Int64")

    # numeriske kolonner
    num_cols = ["Saldo i fjor","Foreløpig Saldo i år",
                "Korreksjon i år","Saldo i år"]
    df[num_cols] = df[num_cols].apply(lambda s: pd.to_numeric(s, errors="coerce"))

    # Regnnr = kol H (+ I hvis utfylt)
    h = df["H"].astype(str).str.strip()
    i = df["I"].fillna("").astype(str).str.strip()
    comb = np.where(i == "", h, h + " " + i)
    df["Regnnr"] = [m.group(0) if (m:=re.search(r"\d+", s)) else ""
                    for s in comb]
    df = df[df["Regnnr"] != ""]

    df["Kontonavn"] = df["Kontonavn"].astype(str).str.strip()
    df["NAkonto"]   = df["NAkonto"].apply(_clean_acct)
    df["Revnr"]     = df["Revnr"].astype(str).str.strip()
    return df.drop(columns=["H", "I"])

# -------------------------------------------------- #
#  bygg target + pivoter
# -------------------------------------------------- #
def build_target(src, maestro, naering, revisjon):
    maestro = maestro.copy(); maestro.index = maestro.index.astype(str)
    tgt = (src
           .merge(maestro.rename(columns={"Value": "Regnskapslinje"}),
                  how="left", left_on="Regnnr", right_index=True)
           .merge(naering.rename(columns={"Value": "NAnavn"}),
                  how="left", left_on="NAkonto", right_index=True)
           .merge(revisjon.rename(columns={"Value": "Revområde"}),
                  how="left", left_on="Revnr", right_index=True)
           .assign(Revområde=lambda d: d["Revområde"].fillna("Ingen match")))

    order = ["Konto","Kontonavn","Saldo i fjor","Foreløpig Saldo i år",
             "Korreksjon i år","Saldo i år",
             "Regnnr","Regnskapslinje","NAkonto","NAnavn","Revnr","Revområde"]
    return tgt.reindex(columns=order)

def piv_regnsk(df):
    p = (df.groupby(["Regnnr","Regnskapslinje"], as_index=False)
           [["Saldo i fjor","Saldo i år"]].sum())
    p["Differanse"]   = p["Saldo i år"] - p["Saldo i fjor"]
    p["Differanse %"] = np.where(p["Saldo i fjor"] == 0, np.nan,
                                 p["Differanse"] / p["Saldo i fjor"])
    return p

# --------- NY piv_konto ----------------------------------------------------
def piv_konto(df):
    p = (df.groupby(["Konto","Kontonavn"], as_index=False)
           .agg({"Saldo i fjor": "sum",
                 "Saldo i år"  : "sum",
                 # metadata fra første forekomst
                 "Regnnr"         : "first",
                 "Regnskapslinje" : "first",
                 "NAkonto"        : "first",
                 "NAnavn"         : "first"}))

    p["Endring"]   = p["Saldo i år"] - p["Saldo i fjor"]
    p["Endring %"] = np.where(p["Saldo i fjor"] == 0, np.nan,
                              p["Endring"] / p["Saldo i fjor"])

    cols = ["Konto","Kontonavn",
            "Saldo i fjor","Saldo i år","Endring","Endring %",
            "Regnnr","Regnskapslinje","NAkonto","NAnavn"]
    return p[cols]
# ---------------------------------------------------------------------------

# -------------------------------------------------- #
#  Excel-hjelpere
# -------------------------------------------------- #
def fmt_thousands(ws, fr, to, cols):
    for c in cols:
        let = get_column_letter(c)
        for cell in ws[f"{let}{fr}:{let}{to}"]:
            cell[0].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

def fmt_percent(ws, fr, to, col):
    let = get_column_letter(col)
    for cell in ws[f"{let}{fr}:{let}{to}"]:
        cell[0].number_format = "0.0%"

def autosize(ws):
    for col in ws.columns:
        w = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = w

# -------------------------------------------------- #
#  vesentlighetsgrense
# -------------------------------------------------- #
def ves_grense(wb, år):
    opp = wb["Oppstilling"]
    def grab(nr: int):
        for row in opp.iter_rows(min_row=2, values_only=True):
            if row[0] == nr:
                return float(row[3]) if row[3] not in (None,"") else np.nan
        return np.nan
    di   = grab(19); dk = grab(79)
    brut = di - dk if not np.isnan(di) and not np.isnan(dk) else np.nan
    rfs  = grab(160); ei = grab(665); ek = grab(715)

    rows = [("Driftsinntekter",       di,  0.01, 0.02),
            ("Bruttofortjeneste",    brut, 0.015,0.03),
            ("Resultat før skatt",    rfs, 0.05, 0.10),
            ("Sum eiendeler",         ei,  0.005,0.01),
            ("Egenkapital",           ek,  0.01, 0.05)]

    ws = wb.create_sheet("Vesentlighetsgrense")
    ws.append(["Periode", år]); ws.append([])
    ws.append(["Type","Beløp","Fra %","Til %","Fra","Til","gj. snitt"])
    for t,b,p1,p2 in rows:
        fra = b*p1 if not np.isnan(b) else ""
        til = b*p2 if not np.isnan(b) else ""
        gsn = (fra+til)/2 if fra!="" and til!="" else ""
        ws.append([t,b,p1,p2,fra,til,gsn])

    last = ws.max_row
    fmt_thousands(ws,4,last,[2,5,6,7])
    fmt_percent(ws,4,last,3); fmt_percent(ws,4,last,4)
    autosize(ws)

# -------------------------------------------------- #
#  CLI / main
# -------------------------------------------------- #
def cli(argv: Optional[List[str]] = None):
    p = argparse.ArgumentParser()
    p.add_argument("base_dir", nargs="?", default=BASE_DIR_DEF, type=Path)
    p.add_argument("--debug", action="store_true")
    p.add_argument("--maestro-sheet", default="Sheet1")
    return p.parse_args(argv)

def main(argv: Optional[List[str]] = None):
    args = cli(argv)
    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO,
                        format="%(levelname)s: %(message)s")

    base = args.base_dir.expanduser().resolve()
    src  = choose_source(base)

    # referanser
    try:
        maestro = read_ref(base/MAESTRO_FILE, key_col=1, val_col=2,
                           sheet=args.maestro_sheet)
    except ValueError:
        logging.warning("Maestro B/C tom – prøver A/B …")
        maestro = read_ref(base/MAESTRO_FILE, key_col=0, val_col=1,
                           sheet=args.maestro_sheet)

    naering  = read_ref(base/NAERING_FILE,  key_col=0, val_col=1)
    revisjon = read_ref(base/REVISJON_FILE, key_col=0, val_col=1)

    src_df = load_source(src)
    tgt_df = build_target(src_df, maestro, naering, revisjon)
    piv_r  = piv_regnsk(tgt_df)
    piv_k  = piv_konto(tgt_df)

    # filnavn
    stem, ext = src.stem, src.suffix
    out_dir   = src.parent
    fname     = f"{stem} Start{ext}"
    v = 2
    while (out_dir/fname).exists():
        fname = f"{stem} Start.v{v}{ext}"; v += 1
    out = out_dir/fname
    logging.info("Lagrer %s", out)

    with pd.ExcelWriter(out, engine="openpyxl") as xls:
        tgt_df.to_excel(xls, "Data",       index=False)
        piv_r .to_excel(xls, "Pivot",      index=False)
        piv_k .to_excel(xls, "KontoPivot", index=False)

        wb = xls.book
        ws = wb["Data"];       fmt_thousands(ws,2,ws.max_row,[3,4,5,6]); autosize(ws)
        ws = wb["Pivot"];      fmt_thousands(ws,2,ws.max_row,[3,4,5]);   fmt_percent(ws,2,ws.max_row,6); autosize(ws)
        ws = wb["KontoPivot"]; fmt_thousands(ws,2,ws.max_row,[3,4,5]);   fmt_percent(ws,2,ws.max_row,6); autosize(ws)

        lag_oppstilling(wb, tgt_df, base/MAESTRO_FILE)
        ves_grense(wb, AAR)

    print("✓ Ferdig:", out)

if __name__ == "__main__":
    main()
