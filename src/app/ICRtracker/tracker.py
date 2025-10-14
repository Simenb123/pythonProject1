# -*- coding: utf-8 -*-
"""
IRC-Tracker: Hent Outlook-eposter, lagre Excel-vedlegg, plukk 'Unique ID',
match mot klientliste (KLIENT_ORGNR) og skriv match_rapport.csv.
Utvidet med opsjon for aksjonærregister-sjekk.

Kjørbar direkte (Run i IDE) ELLER via -m.
"""

# ====== BOOTSTRAP så fila kan kjøres direkte ======
if __name__ == "__main__" and __package__ is None:
    import sys, pathlib
    here = pathlib.Path(__file__).resolve()
    for up in range(2, 6):  # parents[2] .. parents[5]
        cand = here.parents[up] / "src"
        if cand.exists():
            sys.path.insert(0, str(cand))
            __package__ = "app.ICRtracker"
            break
# ==================================================

from __future__ import annotations

import csv
import re
from datetime import datetime, timedelta
from pathlib import Path

import win32com.client  # pywin32
from openpyxl import load_workbook

# --- AR-kobling (valgfri) ---
try:
    from .registry_db import open_db, get_owners, companies_owned_by, normalize_orgnr
    AR_AVAILABLE = True
except Exception:
    AR_AVAILABLE = False

# ========================== KONFIG ==========================
SENDER_EMAIL = "IRC-Norway@no.gt.com"
LOOKBACK_DAYS = 14

DOWNLOAD_DIR = Path(r"F:\Dokument\Kildefiler\irc\irc_vedlegg")
RAPPORT_FIL  = Path(r"F:\Dokument\Kildefiler\irc\rapporter\match_rapport.csv")

KLIENTLISTE  = Path(r"F:\Dokument\Kildefiler\BHL AS klientliste - kopi.xlsx")

REGISTRY_DB_PATH = Path(r"F:\Dokument\Kildefiler\aksjonarregister.db")  # pek til eksisterende .db
RAPPORT_AR_FRA_EPOST = Path(r"F:\Dokument\Kildefiler\irc\rapporter\ar_funn_fra_epost.csv")

SET_OUTLOOK_CATEGORY = True
OUTLOOK_CATEGORY_NAME = "Processed (IRC)"

SHARED_MAILBOX_DISPLAYNAME = ""
# ===========================================================

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def normalize_orgnr_local(val) -> str:
    if AR_AVAILABLE:
        return normalize_orgnr(val)
    if val is None:
        return ""
    return re.sub(r"\D+", "", str(val))

# ---------- klientliste ----------
def read_client_orgs(path: Path) -> set[str]:
    import csv
    from openpyxl import load_workbook
    if not path.exists():
        raise FileNotFoundError(f"Fant ikke klientliste: {path}")
    ext = path.suffix.lower()
    orgs = set()
    if ext in {".xlsx",".xlsm"}:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        ws = wb.active
        header = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        def pick(*alts):
            for a in alts:
                if a in header:
                    return header.index(a)
            return None
        i_org = pick("klient_orgnr","orgnr","organisasjonsnummer")
        if i_org is None:
            raise ValueError("Fant ikke kolonne 'KLIENT_ORGNR'/'orgnr' i klientlista.")
        for row in ws.iter_rows(min_row=2, values_only=True):
            org = normalize_orgnr_local(row[i_org] if len(row)>i_org else "")
            if org:
                orgs.add(org)
        return orgs
    elif ext==".csv":
        def read_csv(encoding:str) -> set[str]:
            s=set()
            with path.open("r",encoding=encoding,newline="") as f:
                r=csv.DictReader(f)
                hdr=[h.lower().strip() for h in (r.fieldnames or [])]
                def pick(*alts):
                    for a in alts:
                        if a in hdr:
                            return a
                    return None
                col = pick("klient_orgnr","orgnr","organisasjonsnummer")
                if not col:
                    raise ValueError("CSV: fant ikke kolonne for orgnr.")
                for row in r:
                    n = normalize_orgnr_local(row.get(col,""))
                    if n:
                        s.add(n)
            return s
        try:
            return read_csv("utf-8-sig")
        except UnicodeDecodeError:
            return read_csv("latin-1")
    else:
        raise ValueError(f"Ukjent klientliste-format: {ext}")

# ---------- Outlook ----------
def get_namespace():
    return win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

def get_root_folder(ns):
    if SHARED_MAILBOX_DISPLAYNAME.strip():
        return ns.Folders[SHARED_MAILBOX_DISPLAYNAME]
    return ns.GetDefaultFolder(6).Parent  # rot for din postboks

def walk_all_mail_folders(root_folder):
    folders=[]
    def rec(f):
        folders.append(f)
        try:
            for i in range(1, f.Folders.Count+1):
                rec(f.Folders.Item(i))
        except Exception:
            pass
    rec(root_folder)
    return folders

def robust_received_filter():
    since = datetime.now() - timedelta(days=LOOKBACK_DAYS)
    return since, since.strftime("%m/%d/%Y %I:%M %p")

def get_smtp_address(msg) -> str:
    try:
        if getattr(msg,"SenderEmailType","")=="EX":
            exu = msg.Sender.GetExchangeUser()
            if exu:
                return (exu.PrimarySmtpAddress or "").lower()
        smtp = (getattr(msg,"SenderEmailAddress","") or "").lower()
        if smtp:
            return smtp
    except Exception:
        pass
    try:
        pa = msg.PropertyAccessor
        prop = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"
        return (pa.GetProperty(prop) or "").lower()
    except Exception:
        return ""

def fetch_messages_from_sender(sender_email: str):
    sender_email = sender_email.lower()
    ns = get_namespace()
    root = get_root_folder(ns)
    folders = walk_all_mail_folders(root)
    since, since_str = robust_received_filter()
    matched=[]
    for f in folders:
        try:
            items=f.Items
            items.Sort("[ReceivedTime]", True)
            try:
                candidates = items.Restrict(f"[ReceivedTime] >= '{since_str}'")
            except Exception:
                candidates = [m for m in items if getattr(m,"ReceivedTime",datetime.min) >= since]
            for m in candidates:
                try:
                    if get_smtp_address(m) == sender_email:
                        matched.append(m)
                except Exception:
                    continue
        except Exception:
            continue
    return matched

def save_excel_attachments(msg, dest_dir: Path):
    saved=[]
    atts = getattr(msg,"Attachments",None)
    if not atts:
        return saved
    for i in range(1, atts.Count+1):
        att = atts.Item(i)
        name=str(att.FileName)
        if name.lower().endswith((".xlsx",".xlsm")):
            try:
                path = dest_dir / f"{msg.EntryID[:8]}_{name}"
                att.SaveAsFile(str(path))
                saved.append(path)
            except Exception as e:
                print(f"  ! Kunne ikke lagre vedlegg '{name}': {e}")
    return saved

def mark_processed(msg):
    if not SET_OUTLOOK_CATEGORY:
        return
    try:
        cats=[c.strip() for c in (msg.Categories or "").split(",") if c.strip()]
        if OUTLOOK_CATEGORY_NAME not in cats:
            cats.append(OUTLOOK_CATEGORY_NAME)
            msg.Categories=", ".join(cats)
            msg.Save()
    except Exception:
        pass

# ---------- Excel 'Unique ID' ----------
def find_unique_id_col(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        if not any(row):
            continue
        for j, val in enumerate(row, start=1):
            if val is None:
                continue
            if str(val).strip().lower() == "unique id":
                return i, j
    return None, None

def extract_unique_ids(xlsx_path: Path):
    ids=[]
    wb = load_workbook(filename=xlsx_path, data_only=True)
    for ws in wb.worksheets:
        header_row, col = find_unique_id_col(ws)
        if not col:
            continue
        r = header_row + 1
        empty = 0
        while True:
            val = ws.cell(row=r, column=col).value
            if val is None or str(val).strip()=="":
                empty += 1
                if empty >= 3:
                    break
            else:
                empty = 0
                ids.append(normalize_orgnr_local(val))
            r += 1
        if ids:
            break
    return [x for x in ids if x]

# ---------- MAIN ----------
def main():
    ensure_dir(DOWNLOAD_DIR)
    ensure_dir(RAPPORT_FIL.parent)

    client_orgs = read_client_orgs(KLIENTLISTE)
    print(f"Klientliste: {len(client_orgs)} orgnr lastet fra {KLIENTLISTE}")

    msgs = fetch_messages_from_sender(SENDER_EMAIL)
    print(f"Fant {len(msgs)} meldinger fra {SENDER_EMAIL} siste {LOOKBACK_DAYS} dager.")

    rows_basic=[]
    rows_ar=[]
    total_saved=0

    ar_conn=None
    if AR_AVAILABLE and REGISTRY_DB_PATH and REGISTRY_DB_PATH.exists():
        try:
            ar_conn = open_db(REGISTRY_DB_PATH)
            print(f"AR-kobling aktiv: {REGISTRY_DB_PATH}")
        except Exception as e:
            print(f"Kunne ikke åpne AR-db: {e}")

    for msg in msgs:
        subject = getattr(msg, "Subject", "")
        received = getattr(msg, "ReceivedTime", None)
        received_iso = (received.strftime("%Y-%m-%d %H:%M") if received else "")

        saved_files = save_excel_attachments(msg, DOWNLOAD_DIR)
        print(f" - '{subject}' → lagret {len(saved_files)} vedlegg")
        total_saved += len(saved_files)

        matches_found = 0
        for f in saved_files:
            ids = extract_unique_ids(f)
            for org in ids:
                hit = "JA" if org in client_orgs else "NEI"
                if hit == "JA":
                    matches_found += 1
                rows_basic.append({
                    "received": received_iso,
                    "subject": subject,
                    "attachment": f.name,
                    "orgnr": org,
                    "match": hit
                })

                if ar_conn:
                    # eiere av org
                    for r in get_owners(ar_conn, org):
                        rel_hit = (normalize_orgnr_local(r["shareholder_orgnr"]) in client_orgs) if r["shareholder_orgnr"] else False
                        rows_ar.append({
                            "received": received_iso,
                            "subject": subject,
                            "attachment": f.name,
                            "client_orgnr": org,
                            "direction": "owned_by",
                            "related_orgnr": r["shareholder_orgnr"] or "",
                            "related_name": r["shareholder_name"] or "",
                            "related_type": r["shareholder_type"] or "",
                            "stake_percent": r["stake_percent"],
                            "shares": r["shares"],
                            "flag_client_crosshit": "JA" if rel_hit else "NEI"
                        })
                    # selskaper org eier
                    for r in companies_owned_by(ar_conn, org):
                        rel_hit = (normalize_orgnr_local(r["company_orgnr"]) in client_orgs)
                        rows_ar.append({
                            "received": received_iso,
                            "subject": subject,
                            "attachment": f.name,
                            "client_orgnr": org,
                            "direction": "owns",
                            "related_orgnr": r["company_orgnr"],
                            "related_name": (r.get("company_name","") if hasattr(r, "get") else r["company_name"]),
                            "related_type": "company",
                            "stake_percent": r["stake_percent"],
                            "shares": r["shares"],
                            "flag_client_crosshit": "JA" if rel_hit else "NEI"
                        })

        if SET_OUTLOOK_CATEGORY and matches_found > 0:
            mark_processed(msg)

    # Rapporter
    with RAPPORT_FIL.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["received","subject","attachment","orgnr","match"])
        writer.writeheader(); writer.writerows(rows_basic)

    if rows_ar:
        ensure_dir(RAPPORT_AR_FRA_EPOST.parent)
        with RAPPORT_AR_FRA_EPOST.open("w", encoding="utf-8", newline="") as f:
            fieldnames = ["received","subject","attachment","client_orgnr","direction",
                          "related_orgnr","related_name","related_type","stake_percent","shares","flag_client_crosshit"]
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader(); w.writerows(rows_ar)

    print(f"\nFerdig. {len(rows_basic)} linjer skrevet til: {RAPPORT_FIL}")
    if rows_ar:
        print(f"AR-funn fra epost: {len(rows_ar)} linjer -> {RAPPORT_AR_FRA_EPOST}")
    print(f"Vedlegg lagret: {total_saved} stk → {DOWNLOAD_DIR}")

if __name__ == "__main__":
    main()
