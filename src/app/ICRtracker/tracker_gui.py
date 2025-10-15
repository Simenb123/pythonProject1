# -*- coding: utf-8 -*-
"""
tracker_gui.py
----------------
Tkinter-GUI for å kjøre og overvåke IRC-tracker + aksjonærregister.

Nyheter:
- Flere avsendere (kommaseparert/semicolon/linjeskift)
- Ekstra kolonner i "Matcher (vedlegg)": related_name, entity_relationship, entity_role, country, unique_id_type
- Mer informativ e-postvarsling (inkluderer hvilke treff)
- Fuzzy personmatch (DL/SL) med korrekt håndtering av æ/ø/å/ä/ö/ü
- AR-oppslag gjøres kun for 9-sifret orgnr (reduserer støy)
"""
from __future__ import annotations

import csv
import json
import os
import queue
import re
import subprocess
import sys
import threading
import traceback
import unicodedata
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# --- Robust import av prosjektpakker ---
if __name__ == "__main__" and __package__ is None:
    here = Path(__file__).resolve()
    for parent in (here.parent, *here.parents):
        if (parent / "app" / "__init__.py").exists():
            sys.path.insert(0, str(parent))
            break

# Import fra tracker-modulen (Outlook/vedlegg/ID/konstanter)
try:
    import app.ICRtracker.tracker as tracker_mod
    from app.ICRtracker.tracker import (
        ensure_dir, read_client_orgs, fetch_messages_from_sender,
        save_excel_attachments, normalize_orgnr_local,
        SENDER_EMAIL as DEFAULT_SENDER_EMAIL,
        LOOKBACK_DAYS as DEFAULT_LOOKBACK,
        DOWNLOAD_DIR as DEFAULT_DOWNLOAD_DIR,
        KLIENTLISTE as DEFAULT_CLIENTS_PATH,
        RAPPORT_FIL as DEFAULT_REPORT_PATH,
        RAPPORT_AR_FRA_EPOST as DEFAULT_AR_REPORT_PATH,
        REGISTRY_DB_PATH as DEFAULT_REG_DB_PATH,
    )
    TRACKER_OK = True
    TRACKER_ERR = None
except Exception as e:
    TRACKER_OK = False
    TRACKER_ERR = e

# AR-bro (foretrukket), ellers registry_db
AR_AVAILABLE = False
AR_BACKEND = "none"
try:
    from app.ICRtracker.ar_bridge import open_db as ar_open_db, get_owners, companies_owned_by, normalize_orgnr  # type: ignore
    AR_AVAILABLE = True
    AR_BACKEND = "ar_bridge"
except Exception:
    try:
        from app.ICRtracker.registry_db import open_db as ar_open_db, get_owners, companies_owned_by, normalize_orgnr  # type: ignore
        AR_AVAILABLE = True
        AR_BACKEND = "registry_db"
    except Exception:
        AR_AVAILABLE = False
        AR_BACKEND = "none"


# ---------------- Konfigmodell ----------------

def _default_path(p: Path) -> str:
    try:
        return str(p)
    except Exception:
        return ""

@dataclass
class GUIConfig:
    sender_emails: str = DEFAULT_SENDER_EMAIL if TRACKER_OK else "IRC-Norway@no.gt.com"
    lookback_days: int = DEFAULT_LOOKBACK if TRACKER_OK else 14
    download_dir: str = _default_path(DEFAULT_DOWNLOAD_DIR) if TRACKER_OK else ""
    clients_path: str  = _default_path(DEFAULT_CLIENTS_PATH) if TRACKER_OK else ""
    report_path: str   = _default_path(DEFAULT_REPORT_PATH) if TRACKER_OK else ""
    ar_report_path: str= _default_path(DEFAULT_AR_REPORT_PATH) if TRACKER_OK else ""
    registry_db_path: str = _default_path(DEFAULT_REG_DB_PATH) if TRACKER_OK else ""

    ar_enable: bool = True
    ar_up_depth: int = 1
    ar_down_depth: int = 1
    ar_min_stake: float = 0.0
    name_score_min: int = 90

    notify_enable: bool = False
    notify_to: str = ""

    schedule_enable: bool = False
    schedule_time: str = "09:00"

def appdata_config_path() -> Path:
    base = os.getenv("APPDATA")
    if base:
        d = Path(base) / "IRCTracker"
        d.mkdir(parents=True, exist_ok=True)
        return d / "gui_config.json"
    return Path(__file__).with_name("gui_config.json")

def load_config() -> GUIConfig:
    p = appdata_config_path()
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            if "sender_email" in data and "sender_emails" not in data:
                data["sender_emails"] = data.pop("sender_email")
            return GUIConfig(**data)
        except Exception:
            pass
    return GUIConfig()

def save_config(cfg: GUIConfig):
    p = appdata_config_path()
    p.write_text(json.dumps(asdict(cfg), ensure_ascii=False, indent=2), encoding="utf-8")

def shell_quote(s: str) -> str:
    return '"' + s.replace('"', '\\"') + '"'


# ---------------- Outlook e-post ----------------

def send_outlook_mail(to_list: List[str], subject: str, body: str, attachments: Optional[List[Path]] = None) -> str:
    try:
        import win32com.client  # type: ignore
    except Exception as e:
        return f"Outlook COM utilgjengelig: {e}"

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = ";".join(t.strip() for t in to_list if t.strip())
        mail.Subject = subject
        mail.Body = body
        for att in attachments or []:
            try:
                mail.Attachments.Add(str(att))
            except Exception as e:
                mail.Body += f"\n(Obs: kunne ikke legge ved {att.name}: {e})"
        mail.Send()
        return "Epost sendt."
    except Exception as e:
        return f"Feil ved sending: {e}"


# ---------------- Navn-normalisering / fuzzy ----------------

# eksplisitt translitterering av nordiske bokstaver før diakritikk-stripping
TRANSLIT = str.maketrans({
    "Æ":"AE","æ":"ae",
    "Ø":"O", "ø":"o",
    "Å":"AA","å":"aa",
    "Ä":"AE","ä":"ae",
    "Ö":"OE","ö":"oe",
    "Ü":"UE","ü":"ue",
})

def strip_accents_keep_translit(s: str) -> str:
    s = s.translate(TRANSLIT)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s

def norm_name(s: str) -> str:
    if not s:
        return ""
    s = strip_accents_keep_translit(s).lower()
    s = re.sub(r"[^a-z0-9\s\-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def fuzzy_ratio(a: str, b: str) -> int:
    import difflib
    return int(round(difflib.SequenceMatcher(None, a, b).ratio() * 100))


# ---------------- DL/SL (personer) fra klientliste ----------------

def read_client_person_names(path: Path) -> List[str]:
    names: List[str] = []
    if not path.exists():
        return names

    def add(n):
        n = (n or "").strip()
        if n:
            names.append(norm_name(n))

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return names
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ws = wb.active
        header = [(cell.value or "").strip() for cell in ws[1]]
        idx = { (h or "").strip().lower(): i for i,h in enumerate(header) }
        cand = [idx[k] for k in idx if k in {"dl","sl","daglig leder","styrets leder"}]
        if not cand:
            for k,i in idx.items():
                if "daglig" in k or "leder" in k or k in {"dl","sl"}:
                    cand.append(i)
        for row in ws.iter_rows(min_row=2, values_only=True):
            for ci in cand:
                if ci < len(row) and row[ci]:
                    add(row[ci])
        return list({n for n in names if n})

    # CSV
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        fields = [ (c or "").strip() for c in (r.fieldnames or []) ]
        lowmap = {c.lower(): c for c in fields}
        cand_keys = []
        for k in list(lowmap.keys()):
            if k in {"dl","sl","daglig leder","styrets leder"} or "daglig" in k or "leder" in k:
                cand_keys.append(lowmap[k])
        for row in r:
            for k in cand_keys:
                add(row.get(k, ""))
    return list({n for n in names if n})


# ---------------- Les flere kolonner fra vedlegg ----------------

def extract_attachment_rows(xlsx_path: Path) -> List[Dict[str, str]]:
    """
    Leser hele arket og returnerer poster med:
     - unique_id, unique_id_type, related_name, entity_relationship, entity_role, country, client_name
    Finner header-rad ved å lete etter 'Unique ID'.
    """
    rows: List[Dict[str, str]] = []
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return rows

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    for ws in wb.worksheets:
        header_row = None
        header_map: Dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            if not any(row):
                continue
            lower = [ (str(x).strip().lower() if x is not None else "") for x in row ]
            if "unique id" in lower:
                header_row = i
                header_map = { (str(v).strip().lower() if v is not None else ""): j
                               for j,v in enumerate(row, start=1) if v}
                break
        if not header_row:
            continue

        def col(*cands: str) -> Optional[int]:
            for cand in cands:
                cand = cand.lower()
                if cand in header_map:
                    return header_map[cand]
                # også substring-match (noen ganger "Entity Role (optional for upload…)")
                for key, idx in header_map.items():
                    if cand in key:
                        return idx
            return None

        c_uid = col("unique id")
        if not c_uid:
            continue
        c_name = col("name of related entity or individual", "name of related entity", "related entity", "name")
        c_country = col("physical location country", "country")
        c_rel = col("entity relationship")
        c_role = col("entity role")
        c_uidt = col("unique id type")
        c_client = col("client name")

        # hent rader under headeren
        blank = 0
        for r in range(header_row+1, ws.max_row+1):
            vals = []
            for cidx in {c_uid, c_name, c_country, c_rel, c_role, c_uidt, c_client}:
                if cidx:
                    vals.append(ws.cell(row=r, column=cidx).value)
            if all(v in (None,"") for v in vals):
                blank += 1
                if blank >= 5:
                    break
                continue
            blank = 0

            def get(cidx: Optional[int]) -> str:
                if not cidx:
                    return ""
                v = ws.cell(row=r, column=cidx).value
                return (str(v).strip() if v is not None else "")

            rows.append({
                "unique_id": get(c_uid),
                "unique_id_type": get(c_uidt),
                "related_name": get(c_name),
                "entity_relationship": get(c_rel),
                "entity_role": get(c_role),
                "country": get(c_country),
                "client_name": get(c_client),
            })
    return rows


# ---------------- AR-traversering + klassifisering ----------------

COMPANY_SUFFIXES = {
    "AS","ASA","AB","GMBH","LIMITED","LTD","BV","OY","APS","AG",
    "SAS","SA","SARL","SRO","SRL","SPA","S.P.A","NV","LLC","PLC",
    "A/S","KFT","OOO","INC","PTY","PTY LTD","BVBA","S.A.",
}

def looks_like_company_name(name: str) -> bool:
    u = strip_accents_keep_translit((name or "").upper())
    for suf in COMPANY_SUFFIXES:
        if re.search(rf"\b{re.escape(suf)}\b", u):
            return True
    tokens = re.findall(r"[A-Z]{2,}", u)
    return len(tokens) >= 2

def classify_related(related_orgnr: str, related_name: str, given_type: str = "") -> str:
    oid = re.sub(r"\D+", "", related_orgnr or "")
    if len(oid) == 4:
        try:
            yr = int(oid)
            if 1900 <= yr <= 2025:
                return "person"
        except Exception:
            pass
    if 8 <= len(oid) <= 14:
        return "company"
    if not oid and looks_like_company_name(related_name or ""):
        return "foreign_company_unknown_id"
    if given_type:
        return given_type
    nm = (related_name or "").strip()
    if " " in nm and not looks_like_company_name(nm):
        return "person"
    return "foreign_company_unknown_id"

def ar_traverse(conn, base_orgnr: str, up_depth: int, down_depth: int,
                min_stake: float = 0.0) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    seen_up = {normalize_orgnr(base_orgnr)}
    seen_down = {normalize_orgnr(base_orgnr)}

    # opp (eiere)
    frontier = [normalize_orgnr(base_orgnr)]
    for level in range(1, up_depth + 1):
        nf = []
        for org in frontier:
            try:
                owners = get_owners(conn, org)
            except Exception:
                owners = []
            for r in owners:
                pct = r.get("stake_percent")
                if pct is not None and pct < min_stake:
                    continue
                rel_org = r.get("shareholder_orgnr") or ""
                rel_name = r.get("shareholder_name") or ""
                typ = classify_related(rel_org, rel_name, r.get("shareholder_type") or "")
                out.append({
                    "direction": "owned_by",
                    "level": level,
                    "client_orgnr": org,
                    "related_orgnr": rel_org,
                    "related_name": rel_name,
                    "related_type": typ,
                    "stake_percent": pct,
                    "shares": r.get("shares")
                })
                nxt = normalize_orgnr(rel_org)
                if nxt and nxt not in seen_up:
                    seen_up.add(nxt); nf.append(nxt)
        frontier = nf

    # ned (datterselskap)
    frontier = [normalize_orgnr(base_orgnr)]
    for level in range(1, down_depth + 1):
        nf = []
        for org in frontier:
            try:
                childs = companies_owned_by(conn, org)
            except Exception:
                childs = []
            for r in childs:
                pct = r.get("stake_percent")
                if pct is not None and pct < min_stake:
                    continue
                rel_org = r.get("company_orgnr") or ""
                rel_name = r.get("company_name") or ""
                typ = classify_related(rel_org, rel_name, "company")
                out.append({
                    "direction": "owns",
                    "level": level,
                    "client_orgnr": org,
                    "related_orgnr": rel_org,
                    "related_name": rel_name,
                    "related_type": typ,
                    "stake_percent": pct,
                    "shares": r.get("shares")
                })
                nxt = normalize_orgnr(rel_org)
                if nxt and nxt not in seen_down:
                    seen_down.add(nxt); nf.append(nxt)
        frontier = nf

    return out


# ---------------- Pipeline ----------------

@dataclass
class RunResult:
    rows_basic: List[Dict[str, object]]
    rows_ar: List[Dict[str, object]]
    saved_files: List[Path]

def _split_addresses(s: str) -> List[str]:
    parts = re.split(r"[;\,\n]+", s or "")
    return [p.strip() for p in parts if p.strip()]

def run_pipeline(cfg: GUIConfig, logq: queue.Queue) -> RunResult:
    def qlog(m): logq.put(m)

    if not TRACKER_OK:
        qlog(f"FEIL: tracker-moduler kunne ikke importeres: {TRACKER_ERR!r}")
        return RunResult([], [], [])

    tracker_mod.LOOKBACK_DAYS = cfg.lookback_days
    ensure_dir(Path(cfg.download_dir))

    # 1) Klientliste
    try:
        client_orgs = read_client_orgs(Path(cfg.clients_path))
        qlog(f"Klientliste: {len(client_orgs)} orgnr lastet.")
    except Exception as e:
        qlog(f"FEIL: kunne ikke lese klientliste (orgnr): {e}")
        return RunResult([], [], [])

    client_persons = []
    try:
        client_persons = read_client_person_names(Path(cfg.clients_path))
        qlog(f"Klientliste: {len(client_persons)} personnavn (DL/SL) lastet.")
    except Exception as e:
        qlog(f"ADVARSEL: kunne ikke lese DL/SL fra klientlisten: {e}")

    # 2) Eposter fra flere avsendere
    messages_by_id = {}
    for addr in _split_addresses(cfg.sender_emails):
        try:
            try:
                msgs = fetch_messages_from_sender(addr, cfg.lookback_days)  # hvis funksjonen tar 2 param
            except TypeError:
                msgs = fetch_messages_from_sender(addr)                      # bakoverkomp
            for m in msgs:
                messages_by_id[getattr(m, "EntryID", id(m))] = m
        except Exception as e:
            qlog(f"ADVARSEL: kunne ikke hente fra {addr}: {e}")
    msgs = list(messages_by_id.values())
    qlog(f"Fant {len(msgs)} meldinger fra {cfg.sender_emails} (siste {cfg.lookback_days} dager).")

    # 3) AR
    conn = None
    if cfg.ar_enable and AR_AVAILABLE:
        try:
            conn = ar_open_db(Path(cfg.registry_db_path) if cfg.registry_db_path else None)
            qlog(f"AR-backend aktiv ({AR_BACKEND}).")
        except Exception as e:
            qlog(f"ADVARSEL: kunne ikke åpne AR: {e}\n{traceback.format_exc(limit=1)}")
            conn = None
    elif cfg.ar_enable:
        qlog("ADVARSEL: AR ikke tilgjengelig.")

    rows_basic: List[Dict[str, object]] = []
    rows_ar: List[Dict[str, object]] = []
    saved_files: List[Path] = []

    # 4) Prosesser meldinger
    for msg in msgs:
        subject = getattr(msg, "Subject", "")
        received = getattr(msg, "ReceivedTime", None)
        received_iso = (received.strftime("%Y-%m-%d %H:%M") if received else "")

        files = save_excel_attachments(msg, Path(cfg.download_dir))
        saved_files.extend(files)
        qlog(f" - '{subject}' → lagret {len(files)} vedlegg")

        for f in files:
            try:
                recs = extract_attachment_rows(f)
            except Exception as e:
                qlog(f"   ! Kunne ikke lese rader fra {f.name}: {e}")
                recs = []

            for rec in recs:
                raw_uid = rec.get("unique_id", "")
                org = normalize_orgnr_local(raw_uid)
                hit = "JA" if org and (org in client_orgs) else "NEI"

                # GUI-linje
                rows_basic.append({
                    "received": received_iso,
                    "subject": subject,
                    "attachment": f.name,
                    "orgnr": org,
                    "match": hit,
                    "related_name": rec.get("related_name",""),
                    "entity_relationship": rec.get("entity_relationship",""),
                    "entity_role": rec.get("entity_role",""),
                    "country": rec.get("country",""),
                    "unique_id_type": rec.get("unique_id_type",""),
                })

                # AR (kun sannsynlige orgnr – 9 siffer)
                if conn and len(org) == 9:
                    try:
                        rels = ar_traverse(conn, org,
                                           cfg.ar_up_depth, cfg.ar_down_depth,
                                           cfg.ar_min_stake)
                        for r in rels:
                            rel_org = normalize_orgnr(r.get("related_orgnr") or "")
                            rel_name = norm_name(r.get("related_name") or "")
                            crosshit = False
                            if rel_org and rel_org in client_orgs:
                                crosshit = True
                            elif r.get("related_type") == "person" and client_persons:
                                best = 0
                                for n in client_persons:
                                    sc = fuzzy_ratio(rel_name, n)
                                    if sc > best:
                                        best = sc
                                        if best >= cfg.name_score_min:
                                            crosshit = True
                                            break
                                r["name_score"] = best
                            r["flag_client_crosshit"] = "JA" if crosshit else "NEI"
                            r["source_attachment"] = f.name
                            r["received"] = received_iso
                            r["subject"] = subject
                        rows_ar.extend(rels)
                    except Exception as e:
                        qlog(f"   ! AR-feil for {org}: {e}")

    qlog(f"Ferdig: {len(rows_basic)} grunnlinjer, {len(rows_ar)} AR-relasjoner, {len(saved_files)} vedlegg.")
    return RunResult(rows_basic, rows_ar, saved_files)


# ---------------- GUI-komponenter ----------------

class Table(ttk.Treeview):
    def __init__(self, master, columns: List[str], **kw):
        super().__init__(master, columns=columns, show="headings", **kw)
        self._columns = columns
        for c in columns:
            self.heading(c, text=c)
            width = 120
            if c in {"subject","related_name"}: width = 260
            if c in {"entity_relationship","entity_role"}: width = 180
            if c in {"received","attachment","orgnr","match","country","unique_id_type"}: width = 120
            self.column(c, width=width, anchor="w")
        vsb = ttk.Scrollbar(master, orient="vertical", command=self.yview)
        hsb = ttk.Scrollbar(master, orient="horizontal", command=self.xview)
        self.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

    def load_rows(self, rows: List[Dict[str, object]]):
        self.delete(*self.get_children())
        for r in rows:
            vals = [r.get(c, "") for c in self._columns]
            self.insert("", "end", values=vals)

    def export_csv(self, dest: Path):
        with dest.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(self._columns)
            for iid in self.get_children():
                w.writerow(list(self.item(iid, "values")))


class TrackerGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IRC-Tracker – Kontrollpanel")
        self.geometry("1400x780")
        self.minsize(1160, 660)

        self.cfg = load_config()
        self._build_ui()

        self._log(f"Tracker-moduler: {'OK' if TRACKER_OK else f'FEIL ({TRACKER_ERR!r})'}")
        self._log(f"AR-backend: {AR_BACKEND} (tilgjengelig={AR_AVAILABLE})")

    def _build_ui(self):
        left = ttk.Frame(self, padding=8); left.pack(side="left", fill="y")
        right = ttk.Frame(self, padding=8); right.pack(side="right", fill="both", expand=True)

        # --- Kjøring/Outlook ---
        box1 = ttk.LabelFrame(left, text="Kjøring (Outlook/Tracker)", padding=8)
        box1.pack(fill="x", pady=4)

        ttk.Label(box1, text="Avsender(e):").grid(row=0, column=0, sticky="w")
        self.e_senders = ttk.Entry(box1, width=42)
        self.e_senders.grid(row=0, column=1, sticky="w", padx=6)
        self.e_senders.insert(0, self.cfg.sender_emails)

        ttk.Label(box1, text="Dager tilbake:").grid(row=1, column=0, sticky="w")
        self.e_lookback = ttk.Spinbox(box1, from_=1, to=60, width=6)
        self.e_lookback.grid(row=1, column=1, sticky="w", padx=6)
        self.e_lookback.delete(0, "end"); self.e_lookback.insert(0, str(self.cfg.lookback_days))

        ttk.Label(box1, text="Nedlastingsmappe:").grid(row=2, column=0, sticky="w")
        self.e_dl = ttk.Entry(box1, width=42); self.e_dl.grid(row=2, column=1, sticky="w", padx=6)
        self.e_dl.insert(0, self.cfg.download_dir)
        ttk.Button(box1, text="…", width=3, command=self._pick_dl).grid(row=2, column=2, padx=2)

        ttk.Label(box1, text="Klientliste (xlsx/csv):").grid(row=3, column=0, sticky="w")
        self.e_clients = ttk.Entry(box1, width=42); self.e_clients.grid(row=3, column=1, sticky="w", padx=6)
        self.e_clients.insert(0, self.cfg.clients_path)
        ttk.Button(box1, text="…", width=3, command=self._pick_clients).grid(row=3, column=2, padx=2)

        ttk.Label(box1, text="Rapport (matcher).csv:").grid(row=4, column=0, sticky="w")
        self.e_report = ttk.Entry(box1, width=42)
        self.e_report.grid(row=4, column=1, sticky="w", padx=6)
        self.e_report.insert(0, self.cfg.report_path)

        ttk.Label(box1, text="Rapport (AR).csv:").grid(row=5, column=0, sticky="w")
        self.e_report_ar = ttk.Entry(box1, width=42)
        self.e_report_ar.grid(row=5, column=1, sticky="w", padx=6)
        self.e_report_ar.insert(0, self.cfg.ar_report_path)

        # --- AR ---
        box2 = ttk.LabelFrame(left, text=f"Aksjonærregister (backend={AR_BACKEND})", padding=8)
        box2.pack(fill="x", pady=6)

        self.v_ar = tk.BooleanVar(value=self.cfg.ar_enable)
        ttk.Checkbutton(box2, text="Slå på AR-relasjoner", variable=self.v_ar).grid(row=0, column=0, sticky="w", columnspan=3)

        ttk.Label(box2, text="Opp-dybde:").grid(row=1, column=0, sticky="e")
        self.e_up = ttk.Spinbox(box2, from_=0, to=5, width=5)
        self.e_up.grid(row=1, column=1, sticky="w", padx=6)
        self.e_up.delete(0, "end"); self.e_up.insert(0, str(self.cfg.ar_up_depth))

        ttk.Label(box2, text="Ned-dybde:").grid(row=1, column=2, sticky="e")
        self.e_down = ttk.Spinbox(box2, from_=0, to=5, width=5)
        self.e_down.grid(row=1, column=3, sticky="w", padx=6)
        self.e_down.delete(0, "end"); self.e_down.insert(0, str(self.cfg.ar_down_depth))

        ttk.Label(box2, text="Min eierandel (%):").grid(row=2, column=0, sticky="e")
        self.e_minpct = ttk.Spinbox(box2, from_=0, to=100, increment=0.5, width=6)
        self.e_minpct.grid(row=2, column=1, sticky="w", padx=6)
        self.e_minpct.delete(0, "end"); self.e_minpct.insert(0, str(self.cfg.ar_min_stake))

        ttk.Label(box2, text="Min navnescore (%):").grid(row=2, column=2, sticky="e")
        self.e_namescore = ttk.Spinbox(box2, from_=50, to=100, increment=1, width=6)
        self.e_namescore.grid(row=2, column=3, sticky="w", padx=6)
        self.e_namescore.delete(0, "end"); self.e_namescore.insert(0, str(self.cfg.name_score_min))

        ttk.Label(box2, text="registry_db (fallback-sti):").grid(row=3, column=0, sticky="w", columnspan=2)
        self.e_regdb = ttk.Entry(box2, width=42)
        self.e_regdb.grid(row=3, column=2, sticky="w", padx=6, columnspan=2)
        self.e_regdb.insert(0, self.cfg.registry_db_path)

        # --- Varsling ---
        box3 = ttk.LabelFrame(left, text="Varsling (Outlook e-post)", padding=8)
        box3.pack(fill="x", pady=6)

        self.v_notify = tk.BooleanVar(value=self.cfg.notify_enable)
        ttk.Checkbutton(box3, text="Send e-post ved treff", variable=self.v_notify).grid(row=0, column=0, sticky="w", columnspan=3)
        ttk.Label(box3, text="Til (;-separert):").grid(row=1, column=0, sticky="e")
        self.e_notify = ttk.Entry(box3, width=42)
        self.e_notify.grid(row=1, column=1, sticky="w", padx=6, columnspan=2)
        self.e_notify.insert(0, self.cfg.notify_to)

        # --- Planlegging ---
        box4 = ttk.LabelFrame(left, text="Planlegging (Task Scheduler)", padding=8)
        box4.pack(fill="x", pady=6)

        self.v_sched = tk.BooleanVar(value=self.cfg.schedule_enable)
        ttk.Checkbutton(box4, text="Aktiver (opprett planlagt jobb)", variable=self.v_sched).grid(row=0, column=0, sticky="w")
        ttk.Label(box4, text="Tid (HH:MM):").grid(row=0, column=1, sticky="e")
        self.e_time = ttk.Entry(box4, width=8); self.e_time.grid(row=0, column=2, sticky="w", padx=6)
        self.e_time.insert(0, self.cfg.schedule_time)
        ttk.Button(box4, text="Opprett jobb", command=self._create_task).grid(row=0, column=3, padx=4)
        ttk.Button(box4, text="Slett jobb", command=self._delete_task).grid(row=0, column=4)

        # --- Knapper ---
        bar = ttk.Frame(left); bar.pack(fill="x", pady=8)
        ttk.Button(bar, text="Lagre innstillinger", command=self._save_cfg).pack(side="left")
        ttk.Button(bar, text="Kjør nå", command=self._run_now).pack(side="left", padx=6)

        # --- Resultater ---
        nb = ttk.Notebook(right); nb.pack(fill="both", expand=True)

        tab1 = ttk.Frame(nb); nb.add(tab1, text="Matcher (vedlegg)")
        cols_basic = ["received","subject","attachment","orgnr","match",
                      "related_name","entity_relationship","entity_role","country","unique_id_type"]
        self.tbl_basic = Table(tab1, columns=cols_basic); self.tbl_basic.pack(fill="both", expand=True)
        bar1 = ttk.Frame(tab1); bar1.pack(fill="x")
        ttk.Button(bar1, text="Eksporter CSV…", command=lambda: self._export_table(self.tbl_basic, "grunnmatcher.csv")).pack(side="left")

        tab2 = ttk.Frame(nb); nb.add(tab2, text="AR-relasjoner")
        cols_ar = ["received","subject","source_attachment","client_orgnr","direction","level",
                   "related_orgnr","related_name","related_type","stake_percent","shares","name_score","flag_client_crosshit"]
        self.tbl_ar = Table(tab2, columns=cols_ar); self.tbl_ar.pack(fill="both", expand=True)
        bar2 = ttk.Frame(tab2); bar2.pack(fill="x")
        ttk.Button(bar2, text="Eksporter CSV…", command=lambda: self._export_table(self.tbl_ar, "ar_relasjoner.csv")).pack(side="left")

        # Logg
        logf = ttk.LabelFrame(right, text="Logg"); logf.pack(fill="both", expand=False, pady=(6,0))
        self.txt = tk.Text(logf, height=8); self.txt.pack(fill="both", expand=True)

    # ---- Handlere ----
    def _pick_dl(self):
        d = filedialog.askdirectory(title="Velg mappe for vedlegg")
        if d:
            self.e_dl.delete(0, "end"); self.e_dl.insert(0, d)

    def _pick_clients(self):
        p = filedialog.askopenfilename(title="Velg klientliste", filetypes=[("Excel/CSV","*.xlsx;*.xlsm;*.csv"),("Alle","*.*")])
        if p:
            self.e_clients.delete(0, "end"); self.e_clients.insert(0, p)

    def _save_cfg(self):
        cfg = self._cfg_from_ui(); save_config(cfg)
        self._log("Innstillinger lagret.")

    def _cfg_from_ui(self) -> GUIConfig:
        return GUIConfig(
            sender_emails=self.e_senders.get().strip(),
            lookback_days=int(self.e_lookback.get()),
            download_dir=self.e_dl.get().strip(),
            clients_path=self.e_clients.get().strip(),
            report_path=self.e_report.get().strip(),
            ar_report_path=self.e_report_ar.get().strip(),
            registry_db_path=self.e_regdb.get().strip(),
            ar_enable=bool(self.v_ar.get()),
            ar_up_depth=int(self.e_up.get()),
            ar_down_depth=int(self.e_down.get()),
            ar_min_stake=float(self.e_minpct.get()),
            name_score_min=int(self.e_namescore.get()),
            notify_enable=bool(self.v_notify.get()),
            notify_to=self.e_notify.get().strip(),
            schedule_enable=bool(self.v_sched.get()),
            schedule_time=self.e_time.get().strip()
        )

    def _run_now(self):
        cfg = self._cfg_from_ui(); save_config(cfg)
        self.tbl_basic.load_rows([]); self.tbl_ar.load_rows([])
        self._log("Starter kjøring …")

        self._q = queue.Queue()
        self._result = {"v": None}

        def worker():
            try:
                res = run_pipeline(cfg, self._q)
                self._result["v"] = res

                # Varsel-epost (mer beskrivende innhold)
                if cfg.notify_enable:
                    basic_hits = [r for r in res.rows_basic if r.get("match") == "JA"]
                    ar_hits = [r for r in res.rows_ar if r.get("flag_client_crosshit") == "JA"]

                    if basic_hits or ar_hits:
                        to = [x.strip() for x in cfg.notify_to.split(";") if x.strip()]

                        def line(r):
                            return (f"{r.get('received','')} | {r.get('subject','')}\n"
                                    f"  ID={r.get('orgnr','')} | {r.get('related_name','')} | "
                                    f"{r.get('entity_relationship','')} | {r.get('entity_role','')} | {r.get('country','')}")

                        top_basic = "\n".join(line(r) for r in basic_hits[:20]) or "(ingen)"
                        top_ar = "\n".join(
                            f"{r.get('received','')} | {r.get('subject','')}\n"
                            f"  rel={r.get('related_name','')} ({r.get('related_type','')}, org={r.get('related_orgnr','')}) "
                            f"dir={r.get('direction','')}, lvl={r.get('level','')} "
                            f"score={r.get('name_score','')}, xhit={r.get('flag_client_crosshit','')}"
                            for r in ar_hits[:20]
                        ) or "(ingen)"

                        body = (
                            f"Kjørt: {datetime.now():%Y-%m-%d %H:%M}\n"
                            f"Grunnmatcher (orgnr): {len(basic_hits)}\n"
                            f"AR-krysstreff: {len(ar_hits)}\n\n"
                            f"--- Grunnmatcher ---\n{top_basic}\n\n"
                            f"--- AR-krysstreff ---\n{top_ar}\n"
                        )
                        status = send_outlook_mail(to, "IRC-Tracker: treff", body)
                        self._q.put(f"Varsel: {status}")

            except Exception as e:
                self._q.put(f"FEIL i kjøring: {e}\n{traceback.format_exc(limit=1)}")

        threading.Thread(target=worker, daemon=True).start()
        self._pump_queue()

    def _pump_queue(self):
        try:
            while True:
                self._log(self._q.get_nowait())
        except queue.Empty:
            pass

        res = self._result.get("v")
        if res is None:
            self.after(150, self._pump_queue); return

        self.tbl_basic.load_rows(res.rows_basic)
        self.tbl_ar.load_rows(res.rows_ar)
        self._log("Kjøring ferdig.")

    def _export_table(self, table: Table, default_name: str):
        p = filedialog.asksaveasfilename(title="Lagre CSV", defaultextension=".csv",
                                         initialfile=default_name, filetypes=[("CSV","*.csv")])
        if not p: return
        try:
            table.export_csv(Path(p)); messagebox.showinfo("Eksport", f"Lagret: {p}")
        except Exception as e:
            messagebox.showerror("Feil", f"Kunne ikke lagre CSV: {e}")

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt.insert("end", f"[{ts}] {msg}\n"); self.txt.see("end")

    # ---- Task Scheduler ----
    @property
    def _task_name(self) -> str: return "IRCTracker_Daily"

    def _create_task(self):
        cfg = self._cfg_from_ui(); save_config(cfg)
        python = sys.executable
        script_path = Path(__file__).resolve().parent / "runner_cli.py"
        if not script_path.exists():
            # fallback: kjør GUI som script (uten argumenter)
            script_path = Path(__file__).resolve()
        t = cfg.schedule_time.strip()
        if not re.match(r"^\d{2}:\d{2}$", t):
            messagebox.showwarning("Tid", "Bruk HH:MM (24t)."); return
        tr = f'{shell_quote(python)} {shell_quote(str(script_path))}'
        cmd = ["schtasks","/Create","/F","/SC","DAILY","/TN",self._task_name,"/TR",tr,"/ST",t]
        try:
            out = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True)
            self._log(out.strip()); messagebox.showinfo("Planlagt jobb","Opprettet/oppdatert.")
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Feil", f"Kunne ikke opprette jobb:\n{e.output}")

    def _delete_task(self):
        cmd = ["schtasks","/Delete","/F","/TN",self._task_name]
        try:
            out = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True)
            self._log(out.strip()); messagebox.showinfo("Planlagt jobb","Slettet (hvis fantes).")
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Feil", f"Kunne ikke slette jobb:\n{e.output}")


# ---------------- main ----------------

def main():
    app = TrackerGUI()
    app.mainloop()

if __name__ == "__main__":
    main()
