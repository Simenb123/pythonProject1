# -*- coding: utf-8 -*-
"""
matcher.py
Relasjonslogikk mellom klientliste og aksjonærregister.

- Leser klientliste (xlsx/csv) -> [{'orgnr','navn'}]
- Finner direkte og indirekte relasjoner via registry_db ELLER db_compat_adapter
- Fuzzy navn med rapidfuzz (hvis tilgjengelig) eller difflib
"""
from __future__ import annotations

import csv
import difflib
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

# --- AR-kobling (adapter først, fallback til standard) ---
try:
    from .db_compat_adapter import (
        open_db, get_owners, companies_owned_by, get_company_name, search_name_candidates,
        normalize_orgnr, normalize_name,
    )
    USING_ADAPTER = True
    # print("matcher.py: bruker db_compat_adapter")
except Exception:
    from .registry_db import (
        open_db, get_owners, companies_owned_by, get_company_name, search_name_candidates,
        normalize_orgnr, normalize_name,
    )
    USING_ADAPTER = False
    # print("matcher.py: bruker registry_db")

try:
    from rapidfuzz import fuzz
    def _score(a: str, b: str) -> int:
        return int(fuzz.token_sort_ratio(a, b))
except Exception:
    def _score(a: str, b: str) -> int:
        return int(difflib.SequenceMatcher(None, a.lower(), b.lower()).ratio() * 100)

# ---------- Klientliste ----------
def load_clients(path: Path) -> List[Dict[str, str]]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        ws = wb.active
        header = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        def pick(*alts):
            for a in alts:
                if a in header:
                    return header.index(a)
            return None
        i_org = pick("klient_orgnr", "orgnr", "organisasjonsnummer")
        i_navn = pick("klient_navn", "navn", "klientnavn", "company_name")
        if i_org is None:
            raise ValueError("Fant ikke kolonne for orgnr i klientlista.")
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            org = normalize_orgnr(row[i_org] if len(row) > i_org else "")
            navn = normalize_name(row[i_navn] if (i_navn is not None and len(row) > i_navn) else "")
            if org:
                out.append({"orgnr": org, "navn": navn})
        return out

    elif ext == ".csv":
        def read_csv(encoding: str) -> List[Dict[str, str]]:
            out = []
            with path.open("r", encoding=encoding, newline="") as f:
                r = csv.DictReader(f)
                hdr = [h.lower().strip() for h in (r.fieldnames or [])]
                def pick(*alts):
                    for a in alts:
                        if a in hdr:
                            return a
                    return None
                col_org = pick("klient_orgnr", "orgnr", "organisasjonsnummer")
                col_name = pick("klient_navn", "navn", "company_name")
                if not col_org:
                    raise ValueError("CSV: fant ikke kolonne for orgnr.")
                for row in r:
                    org = normalize_orgnr(row.get(col_org, ""))
                    name = normalize_name(row.get(col_name, ""))
                    if org:
                        out.append({"orgnr": org, "navn": name})
            return out
        try:
            return read_csv("utf-8-sig")
        except UnicodeDecodeError:
            return read_csv("latin-1")
    else:
        raise ValueError(f"Ukjent klientliste-format: {ext}")

# ---------- Fuzzy navn ----------
def fuzzy_match_name_against_registry(conn, name: str, min_score: int = 90, limit: int = 20) -> List[Tuple[str, str, int]]:
    name_n = normalize_name(name)
    if not name_n:
        return []
    candidates = search_name_candidates(conn, name_n, limit=limit * 5)
    ranked = [(n, o, _score(name_n, n)) for (n, o) in candidates]
    ranked.sort(key=lambda x: x[2], reverse=True)
    return [x for x in ranked if x[2] >= min_score][:limit]

# ---------- Relasjoner ----------
def relations_for_client(conn, client_orgnr: str, client_name: str, client_orgs_set: set[str],
                         min_name_score: int = 90) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    org = normalize_orgnr(client_orgnr)

    # 1) Eiere av klienten
    owners = get_owners(conn, org)
    for r in owners:
        rel = {
            "client_orgnr": org,
            "client_name": client_name,
            "direction": "owned_by",
            "related_name": r["shareholder_name"],
            "related_orgnr": r["shareholder_orgnr"] or "",
            "related_type": r["shareholder_type"] or "",
            "stake_percent": r["stake_percent"],
            "shares": r["shares"],
            "company_orgnr": r["company_orgnr"],  # = org
            "company_name": (get_company_name(conn, org) or "")
        }
        rel["flag_client_crosshit"] = (normalize_orgnr(rel["related_orgnr"]) in client_orgs_set) if rel["related_orgnr"] else False
        out.append(rel)

    # 2) Selskaper klienten eier
    owned = companies_owned_by(conn, org)
    for r in owned:
        rel = {
            "client_orgnr": org,
            "client_name": client_name,
            "direction": "owns",
            "related_name": r["company_name"] or "",
            "related_orgnr": r["company_orgnr"],
            "related_type": "company",
            "stake_percent": r["stake_percent"],
            "shares": r["shares"],
            "company_orgnr": r["company_orgnr"],
            "company_name": r["company_name"] or ""
        }
        rel["flag_client_crosshit"] = (normalize_orgnr(rel["related_orgnr"]) in client_orgs_set)
        out.append(rel)

    # 3) Fuzzy navn (dersom ingen strukturelle treff)
    if client_name and not owners and not owned:
        for (cand_name, cand_org, score) in fuzzy_match_name_against_registry(conn, client_name, min_score=min_name_score):
            rel = {
                "client_orgnr": org,
                "client_name": client_name,
                "direction": "name_fuzzy_hit",
                "related_name": cand_name,
                "related_orgnr": cand_org,
                "related_type": "company",
                "stake_percent": None,
                "shares": None,
                "company_orgnr": cand_org,
                "company_name": cand_name,
                "fuzzy_score": score,
                "flag_client_crosshit": (normalize_orgnr(cand_org) in client_orgs_set)
            }
            out.append(rel)

    return out

def scan_all_clients(conn, clients: List[Dict[str, str]], min_name_score: int = 90) -> List[Dict[str, object]]:
    orgset = {c["orgnr"] for c in clients}
    all_rows: List[Dict[str, object]] = []
    for c in clients:
        all_rows.extend(relations_for_client(conn, c["orgnr"], c["navn"], orgset, min_name_score))
    return all_rows
