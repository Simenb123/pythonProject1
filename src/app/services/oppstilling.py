# oppstilling.py – 2025-05-24
"""
Legger arket «Oppstilling» (resultat- & balanse-oppstilling + nøkkeltall)
i en openpyxl-workbook.
"""
from __future__ import annotations
import logging, re
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# ――― små hjelpere ―――
def _clean(n):                        # kontonr → str uten .0
    if pd.isna(n): return ""
    if isinstance(n, (int, np.integer)): return str(n)
    if isinstance(n, float) and n.is_integer(): return str(int(n))
    return str(n).strip()

_pct   = lambda a, b: 0.0 if b == 0 else a / b
_ratio = lambda a, b: 0.0 if b == 0 else a / b

# ――― 1) Maestro-def ―――
def _load_defs(path: Path):
    xl = pd.ExcelFile(path, engine="openpyxl")
    s2 = xl.parse("Sheet2", header=None, usecols="A,B,P,R")
    s2.columns = ["nr","navn","sign","snu"]; s2["nr"] = s2["nr"].apply(_clean)
    inter = xl.parse("Intervall", header=None, usecols="A:H", skiprows=1)
    inter.columns = ["lo","form","hi","sumNr","sumNavn","sign","lvl","_"]
    inter["form"] = inter["form"].astype(str).str.lower().str.strip()
    for c in ("lo","hi","sumNr"):
        inter[c] = pd.to_numeric(inter[c], errors="coerce")
    inter = inter.dropna(subset=["sumNr"])
    inter["lo"]  = inter["lo"].fillna(0).astype(int).astype(str)
    inter["hi"]  = inter["hi"].fillna(0).astype(int).astype(str)
    inter["sumNr"]= inter["sumNr"].astype(int).astype(str)
    return s2, inter[["lo","form","hi","sumNr","sumNavn","sign","lvl"]]

# ――― 2) detaljverdier + fortegn ―――
def _detail_dict(df: pd.DataFrame):
    grp = df.groupby("Regnnr")[["Saldo i fjor","Saldo i år"]].sum().reset_index()
    return {_clean(r["Regnnr"]): (r["Saldo i fjor"], r["Saldo i år"])
            for _,r in grp.iterrows()}

def _apply_neg(detail, sheet2):
    txt = sheet2["sign"].astype(str).str.lower().str.strip()
    neg_num = pd.to_numeric(sheet2["sign"], errors="coerce") < 0
    neg_set = set(sheet2.loc[
        txt.str.contains("neg") | txt.str.contains("-") | neg_num, "nr"])
    for nr in neg_set & detail.keys():
        c,d = detail[nr]; detail[nr] = (-c,-d)

def _collect_snu(sheet2):              # linjer merket «snu»
    return set(sheet2.loc[
        sheet2["snu"].astype(str).str.lower().str.strip().eq("snu"), "nr"])

# ――― 3) rekursiv sum-motor ―――
def _sum_engine(detail, inter, snu_set):
    ftab = {r["sumNr"]: (r["form"], r["lo"], r["hi"],
                         str(r["sign"]).lower().strip())
            for _,r in inter.iterrows()}
    cache: dict[str,tuple[float,float]] = {}
    def get_detail(n:str):
        c,d = detail.get(n,(0.0,0.0))
        if n in snu_set: c,d = -c,-d
        return c,d
    def get_sum(n:str):
        if n in cache: return cache[n]
        if n not in ftab: return get_detail(n)
        form,lo,hi,sign = ftab[n]
        if form=="intervall":
            c=d=0.0
            for i in range(int(lo), int(hi)+1):
                ci,di = get_detail(str(i)); c+=ci; d+=di
        else:
            ac,ad = get_sum(lo); bc,bd = get_sum(hi)
            c,d = (ac+bc, ad+bd) if form=="pluss" else (ac-bc, ad-bd)
        if sign=="negativ": c,d = -c,-d
        cache[n]=(c,d); return c,d
    return get_sum

def _calc_sums(detail, inter, snu):
    gs = _sum_engine(detail, inter, snu)
    return {r["sumNr"]: (*gs(r["sumNr"]), r["sumNavn"], int(r["lvl"]))
            for _,r in inter.iterrows()}

# ――― 4) nøkkeltall inkl. kontrollsummer ―――
def _keys(ds):
    g = lambda k: ds.get(k,(0.0,0.0))[0]
    h = lambda k: ds.get(k,(0.0,0.0))[1]
    out={}
    brut_c,brut_d = g("10")-g("20"), h("10")-h("20")
    out["Brutto"]=(brut_c,brut_d)
    out["Brutto%"]=(_pct(brut_c,g("10")), _pct(brut_d,h("10")))
    dr_c,dr_d = g("19"),h("19")
    out["Driftsmargin"]=(_pct(g("80"),dr_c), _pct(h("80"),dr_d))
    out["NettoRes"]=(_pct(g("280"),dr_c), _pct(h("280"),dr_d))
    out["Kundefordr%"]=(_pct(g("610"),dr_c), _pct(h("610"),dr_d))
    out["Lonn%"]=(_pct(g("40"),dr_c), _pct(h("40"),dr_d))
    out["AnnenDrift%"]=(_pct(g("70"),dr_c), _pct(h("70"),dr_d))
    out["RFS%"]=(_pct(g("160"),dr_c), _pct(h("160"),dr_d))
    e_c = dr_c-g("79")+g("50"); e_d = dr_d-h("79")+h("50")
    out["EBITDA"]=(e_c,e_d)
    out["EBITDA%"]=(_pct(e_c,dr_c), _pct(e_d,dr_d))
    out["Varelager%"]=(_pct(g("605"),g("20")), _pct(h("605"),h("20")))
    out["Avskriv%"]=(_pct(g("50"),g("555")), _pct(h("50"),h("555")))
    out["Arbeidskap"]=(g("660")-g("810"), h("660")-h("810"))
    out["EkAndel"]=(_pct(g("715"),g("665")), _pct(h("715"),h("665")))
    out["Gjeldsgrad"]=(_ratio(g("820"),g("715")), _ratio(h("820"),h("715")))
    long_c,long_d = g("735")+g("760"), h("735")+h("760")
    out["LangGjeld%"]=(_pct(long_c,g("820")), _pct(long_d,h("820")))
    out["Likvid1"]=(_ratio(g("660"),g("810")), _ratio(h("660"),h("810")))
    out["Likvid2"]=(_ratio(g("660")-g("605"),g("810")),
                    _ratio(h("660")-h("605"),h("810")))
    out["Likvid3"]=(_ratio(g("655"),g("810")), _ratio(h("655"),h("810")))
    out["LevGjeld%"]=(_pct(g("780"),g("20")+g("70")),
                      _pct(h("780"),h("20")+h("70")))
    # kontroll
    out["BalDiff"] = (g("850")-g("665"), h("850")-h("665"))
    out["Udisponert"] = (g("280")-g("350"), h("280")-h("350"))
    return out

# ――― 5) formel-tekst med navn ―――
_num_pat = re.compile(r"#(\d+)")
def _with_names(raw:str, names:dict[str,str]) -> str:
    return _num_pat.sub(lambda m: f"#{m.group(1)} {names.get(m.group(1),'')}".strip(), raw)

# ――― 6) layout-konstanter ―――
_HEAD = PatternFill("solid", fgColor="DCE6F1")
_HEAD_FONT = Font(bold=True)
_TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
_SUM_FILL={1:"E6E6E6",2:"C8C8C8",3:"787878"}

# ――― 7) skriv hovedtabell med to seksjoner ―――
def _write_table(ws, detail, sums, names):
    rows=[(int(n),names.get(n,n),c,d,False,0) for n,(c,d) in detail.items()]
    rows+=[(int(n),nav,c,d,True,lvl) for n,(c,d,nav,lvl) in sums.items()]
    rows.sort(key=lambda r:r[0])

    res_rows = [r for r in rows if r[0] < 530]     # result til og med nr 350
    bal_rows = [r for r in rows if r[0] >= 530]

    cur = 1
    def _add_title(txt):
        nonlocal cur
        ws.append(["",txt]); cur += 1
        cell = ws[cur-1][1]
        cell.fill=_TITLE_FILL; cell.font=Font(bold=True,size=12)
        ws.merge_cells(start_row=cur-1,start_column=2,
                       end_row=cur-1,end_column=7)

    def _add_header():
        nonlocal cur
        ws.append(["nr","Regnskapslinje","Saldo i fjor","Saldo i år",
                   "Endring","Endr i %",""]); cur += 1
        for c in ws[cur-1]:
            c.fill=_HEAD; c.font=_HEAD_FONT
            c.alignment=Alignment(horizontal="center")

    def _add_rows(row_list):
        nonlocal cur
        for nr,nav,c,d,is_sum,lvl in row_list:
            diff,dp = d-c,(0.0 if c==0 else (d-c)/abs(c))
            ws.append([nr,nav,c,d,diff,dp,""]); cur += 1
            if is_sum:
                fill=PatternFill("solid",fgColor=_SUM_FILL.get(lvl,"D2D2D2"))
                for cl in ws[cur-1]:
                    cl.fill=fill; cl.font=Font(bold=True,
                            color="000000" if lvl<3 else "FFFFFF")

    # resultat-del
    _add_title("RESULTATREGNSKAP")
    _add_header()
    _add_rows(res_rows)

    # én blank rad
    ws.append([]); cur += 1

    # balanse-del
    _add_title("BALANSE")
    _add_header()
    _add_rows(bal_rows)

    # tallformater
    for col in ("C","D","E"):
        for cl in ws[f"{col}1":f"{col}{cur}"]:
            cl[0].number_format="#,##0"
    for cl in ws[f"F1":f"F{cur}"]: cl[0].number_format="0.0%"

    thin=Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws[2:cur]:
        for cl in row: cl.border=thin
    return cur  # siste rad-nr

# ――― 8) skriv nøkkeltall (to blanke rader før) ―――
def _write_keys(ws, start:int, keys, names):
    defs=[("Bruttofortjeneste","Brutto",False,"#10 - #20"),
          ("Bruttofortjeneste (%)","Brutto%",True,"(#10-#20)/#10"),
          ("Driftsmargin (%)","Driftsmargin",True,"#80 / #19"),
          ("Nettoresultatmargin (%)","NettoRes",True,"#280 / #19"),
          ("Likviditetsgrad 1","Likvid1",False,"#660 / #810"),
          ("Likviditetsgrad 2","Likvid2",False,"(#660-#605)/#810"),
          ("Likviditetsgrad 3","Likvid3",False,"#655 / #810"),
          ("Egenkapitalandel (%)","EkAndel",True,"#715 / #665"),
          ("Gjeldsgrad","Gjeldsgrad",False,"#820 / #715"),
          ("Kundefordringer i % av sum driftsinntekter","Kundefordr%",True,"#610 / #19"),
          ("Varelager i % av Varekostnad","Varelager%",True,"#605 / #20"),
          ("Langsiktig gjeld i % av sum gjeld","LangGjeld%",True,"(#735+#760)/#820"),
          ("Avskrivninger i % av varige driftsmidler","Avskriv%",True,"#50 / #555"),
          ("Arbeidskapital","Arbeidskap",False,"#660 - #810"),
          ("Lønnskostnad i % av sum driftsinntekter","Lonn%",True,"#40 / #19"),
          ("Annen driftskostnad i % av sum driftsinntekter","AnnenDrift%",True,"#70 / #19"),
          ("EBITDA","EBITDA",False,"#19 - (#79 - #50)"),
          ("EBITDA (%)","EBITDA%",True,"EBITDA / #19"),
          ("Resultat før skattekostnad (%) av sum driftsinntekter","RFS%",True,"#160 / #19"),
          ("Leverandørgjeld i % av vare- og driftskost","LevGjeld%",True,"#780/(#20+#70)"),
          ("Balansedifferanse","BalDiff",False,"#850 - #665"),
          ("Udisponert resultat","Udisponert",False,"#280 - #350")
    ]
    # to blanke rader
    ws.append([]); ws.append([])
    start += 2

    ws.append(["","NØKKELTALL","Saldo i fjor","Saldo i år",
               "Endring","Endring i %","Formel"]); start += 1
    hdr=ws[start]
    hdr[1].fill=_HEAD
    for cl in hdr[1:]: cl.font=_HEAD_FONT
    ws.merge_cells(start_row=start,start_column=2,end_row=start,end_column=2)

    row=start+1
    for i,(title,key,is_pct,formula) in enumerate(defs,1):
        c_val,d_val = keys.get(key,(0.0,0.0))
        diff=d_val-c_val; diff_pct=0.0 if c_val==0 else diff/abs(c_val)
        ws.append([i,title,c_val,d_val,diff,diff_pct,
                   _with_names(formula,names)])
        fmt="0.0%" if is_pct else "#,##0"
        for cl in ws[f"C{row}":f"D{row}"][0]: cl.number_format=fmt
        ws[f"E{row}"].number_format=fmt; ws[f"F{row}"].number_format="0.0%"
        row += 1

    for col in range(1,8):
        letter=get_column_letter(col)
        ws.column_dimensions[letter].width=max(
            len(str(cl.value or "")) for cl in ws[letter])+2

# ――― 9) publik funksjon ―――
def lag_oppstilling(workbook, df_data: pd.DataFrame, maestro_path: Path):
    logging.info("Lager regnskapsoppstilling med nøkkeltall …")
    sheet2, inter = _load_defs(maestro_path)
    names = dict(zip(sheet2["nr"], sheet2["navn"]))

    detail = _detail_dict(df_data)
    _apply_neg(detail, sheet2)
    snu = _collect_snu(sheet2)

    sums = _calc_sums(detail, inter, snu)
    combined = {**detail, **{k:(c,d) for k,(c,d,_,_) in sums.items()}}
    keys = _keys(combined)

    ws = workbook.create_sheet("Oppstilling")
    last = _write_table(ws, detail, sums, names)
    _write_keys(ws, last, keys, names)

    ws.freeze_panes = ws["A5"]           # etter overskriftene
    logging.info("Oppstilling lagt til.")
