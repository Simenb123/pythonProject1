# utvalg_logikk.py – 2025-06-07 (r3 – saldo & HB valgte kontoer)
from __future__ import annotations
from pathlib import Path
from typing import Mapping, List, Tuple
import math, datetime as dt, re, unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference

# -------- (hjelper-funksjoner _norm, _auto_fmt er som før) -------------
#  ...  kopiert uendret fra din eksisterende versjon ...

# ═══════════  hoved-funksjon  ═════════════════════════════════════════
def kjør_bilagsuttrekk(
    src: Path,
    konto_rng: Tuple[int, int],
    belop_intervaller: List[Tuple[float, float]],
    n_bilag: int,
    *,
    meta: Mapping[str, any],
) -> dict[str, Path]:

    df_full = _apply(_les(src, meta), meta)

    # ny parameter: valgte kontoliste
    kontoliste = meta.get("kontoliste", [])
    if kontoliste:
        df_pop = df_full[df_full["konto"].isin(kontoliste)]
    else:
        df_pop = df_full[df_full["konto"].between(*konto_rng)]

    # beløpsfilter
    if belop_intervaller:
        mask = pd.Series(False, index=df_pop.index)
        for lo, hi in belop_intervaller:
            mask |= df_pop["beløp"].between(lo, hi)
        df_pop = df_pop[mask]

    # trekk bilag
    bilags = df_pop["bilagsnr"].drop_duplicates()
    if n_bilag > len(bilags):
        raise ValueError(f"Pop har kun {len(bilags)} bilag")
    valgte = bilags.sample(n=n_bilag).sort_values()

    df_utvalg = df_pop[df_pop["bilagsnr"].isin(valgte)]
    df_full_utvalg = df_full[df_full["bilagsnr"].isin(valgte)]

    # ---- nye ark ------------------------------------------------------
    saldo_per_konto = df_full_utvalg.groupby("konto")["beløp"]\
                        .sum().reset_index(name="Saldo")

    hb_valgte = df_full[df_full["konto"].isin(kontoliste)] if kontoliste else pd.DataFrame()

    # ---- skriv Excel --------------------------------------------------
    out = src.parent / f"Bilag_uttrekk_{n_bilag}.xlsx"
    i=1
    while out.exists():
        out = out.with_stem(out.stem + f"_v{i}"); i+=1

    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        df_utvalg.to_excel(xw, sheet_name="Kun_intervallet", index=False)
        df_full_utvalg.to_excel(xw, sheet_name="Fullt_bilagsutvalg", index=False)
        saldo_per_konto.to_excel(xw, sheet_name="Saldo_per_konto", index=False)
        if not hb_valgte.empty:
            hb_valgte.to_excel(xw, sheet_name="HB_valgte_kontoer", index=False)

    # enkel auto-format
    wb = load_workbook(out)
    for ws in wb.worksheets:
        _auto_fmt(ws)
    wb.save(out)

    return {"uttrekk": out, "valgte_bilag": valgte.tolist()}
