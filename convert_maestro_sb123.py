# -*- coding: utf-8 -*-
"""
Python‑script som etterligner VBA‑makroen «OmorganisereMaestroSB123».

Versjon 0.4.1 – 09‑05‑2025
-------------------------
* **Bug‑fix:** Rettet skrivefeil som ga `SyntaxError: cannot assign to function call` –
  dobbelt likhet på linjen som leste Revisjons‑filen.
* Ingen andre endringer.
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from tkinter import Tk, filedialog
from typing import Final, List, Optional, Union

import pandas as pd

# ---------------------------------------------------------------------------
# KONFIG
# ---------------------------------------------------------------------------
DEFAULT_BASE_DIR: Path = Path(r"F:/Dokument/2/SB")
MAESTRO_FILE:  Final[str] = "RES OG BAL MAESTRO.xlsx"
NAERING_FILE:  Final[str] = "Naeringsspesifikasjon.xlsx"
REVISJON_FILE: Final[str] = "Revisjonsområder.xlsx"

KUNDENR: str = "0000"
KUNDE:   str = "Demokunde AS"
AAR:     str = "2024"

pd.options.mode.copy_on_write = True

# ---------------------------------------------------------------------------
# Dialog
# ---------------------------------------------------------------------------

def choose_source_file(initial_dir: Path) -> Path:
    logging.info("Åpner filvelger i %s", initial_dir)
    root = Tk(); root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Velg Saldobalanse/Kontoplan fra Maestro",
        initialdir=str(initial_dir),
        filetypes=[("Excel‑filer", "*.xlsx *.xls *.xlsm"), ("Alle filer", "*.*")],
    )
    root.destroy()
    if not file_path:
        sys.exit("Ingen fil valgt – prosessen avbrytes.")
    return Path(file_path)

# ---------------------------------------------------------------------------
# Reference‑loader
# ---------------------------------------------------------------------------

def read_reference_file(
    path: Path,
    *,
    key_col: int = 0,
    value_col: int = 1,
    numeric_keys: bool = False,
    sheet: Union[int, str] = 0,
) -> pd.DataFrame:
    logging.info("Leser referansefil %s (ark=%s, key_col=%d, val_col=%d)", path, sheet, key_col, value_col)

    usecols = [key_col, value_col]
    df = pd.read_excel(path, sheet_name=sheet, usecols=usecols, header=None, engine="openpyxl")
    df.columns = ["Key", "Value"]

    if numeric_keys:
        df["Key"] = pd.to_numeric(df["Key"], errors="coerce")
        df = df[df["Key"].notna()].astype({"Key": "Int64"})
    else:
        df["Key"] = df["Key"].astype(str).str.strip()
        df = df[df["Key"].str.match(r"^\d+$", na=False)]

    if df.empty:
        raise ValueError(f"{path}: fant ingen gyldige rader i kolonnene {key_col}/{value_col}")

    return df.set_index("Key")

# ---------------------------------------------------------------------------
# Source data
# ---------------------------------------------------------------------------

def _clean_account(val) -> str:
    if pd.isna(val):
        return ""
    try:
        return str(int(float(val)))
    except Exception:
        return str(val).strip()


def load_source_data(path: Path) -> pd.DataFrame:
    logging.info("Leser kildefil %s", path)
    src = pd.read_excel(path, sheet_name=0, header=None, skiprows=3, engine="openpyxl")

    colmap = {0: "Konto", 1: "Kontonavn", 6: "Saldo i fjor", 3: "Foreløpig Saldo i år", 4: "Korreksjon i år", 5: "Saldo i år",
              7: "H", 8: "I", 12: "NAkonto", 15: "Revnr"}
    df = src[list(colmap)].rename(columns=colmap)

    for c in ["Konto", "Saldo i fjor", "Foreløpig Saldo i år", "Korreksjon i år", "Saldo i år"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df["Kontonavn"] = df["Kontonavn"].astype(str).str.strip()
    df["Regnnr"] = (df["H"].astype(str).str.strip() + " " + df["I"].astype(str).str.strip()).str.strip()
    df["NAkonto"] = df["NAkonto"].apply(_clean_account)
    df["Revnr"] = df["Revnr"].astype(str).str.strip()

    return df.drop(columns=["H", "I"])

# ---------------------------------------------------------------------------
# Build + pivot
# ---------------------------------------------------------------------------

def build_target(df_src, df_maestro, df_naering, df_rev):
    tgt = (df_src
           .merge(df_maestro.rename(columns={"Value": "Regnskapslinje"}), how="left", left_on="Regnnr", right_index=True)
           .merge(df_naering.rename(columns={"Value": "NAnavn"}), how="left", left_on="NAkonto", right_index=True)
           .merge(df_rev.rename(columns={"Value": "Revområde"}), how="left", left_on="Revnr", right_index=True))
    tgt.loc[tgt["Revområde"].isna(), "Revområde"] = "Ingen match"
    return tgt


def create_pivot(df):
    piv = df.groupby(["Regnnr", "Regnskapslinje"], as_index=False)[["Saldo i fjor", "Saldo i år"]].sum()
    piv["Differanse"] = piv["Saldo i år"] - piv["Saldo i fjor"]
    return piv

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def autosize(ws):
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.iter_cols(values_only=True), 1):
        ws.column_dimensions[get_column_letter(i)].width = max((len(str(v)) if v else 0) for v in col) + 2

def freeze(ws):
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(1)}2"

# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def cli(argv: Optional[List[str]] = None):
    p = argparse.ArgumentParser()
    p.add_argument("base_dir", nargs="?", default=DEFAULT_BASE_DIR, type=Path)
    p.add_argument("--debug", action="store_true")
    p.add_argument("--maestro-sheet", default="Sheet1")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None):
    args = cli(argv)
    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO, format="%(levelname)s: %(message)s")

    base_dir = args.base_dir.expanduser().resolve()
    src_path = choose_source_file(base_dir)

    df_maestro = read_reference_file(base_dir/MAESTRO_FILE, key_col=1, value_col=2, sheet=args.maestro_sheet)
    df_naering = read_reference_file(base_dir/NAERING_FILE)
    df_revisjon = read_reference_file(base_dir/REVISJON_FILE)

    df_source = load_source_data(src_path)
    df_target = build_target(df_source, df_maestro, df_naering, df_revisjon)
    df_pivot = create_pivot(df_target)

    out_path = find_next_versioned_filename(base_dir, f"{KUNDENR} {KUNDE} Saldobalanse {AAR}.xlsx")
    logging.info("Lagrer %s", out_path)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xls:
        df_target.to_excel(xls, sheet_name="Data", index=False)
        df_pivot.to_excel(xls, sheet_name="Pivot", index=False)
        for ws in xls.book.worksheets:
            freeze(ws)
            autosize(ws)

    print(f"✓ Fil lagret: {out_path}")


if __name__ == "__main__":
    main()
